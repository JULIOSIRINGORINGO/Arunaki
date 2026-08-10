import os
import re
import openpyxl
from openpyxl.utils import get_column_letter

def parse_txt_report(txt_path):
    """Parses daily transaction text file and extracts structured data."""
    with open(txt_path, 'r', encoding='utf-8') as f:
        lines = [l.strip() for l in f.readlines() if l.strip()]

    data = {
        'date_day': None,
        'date_month_name': None,
        'date_year': None,
        'pemasukan_items': [],
        'total_pemasukan': 0,
        'total_bca': 0,
        'total_bni': 0,
        'total_bri': 0,
        'total_mandiri': 0,
        'total_cash': 0,
        'total_pengeluaran': 0,
        'pengeluaran_items': []
    }

    current_section = None

    for line in lines:
        # Detect header date
        date_match = re.search(r'(\d{1,2})\s+(AGUSTUS|JANUARI|FEBRUARI|MARET|APRIL|MEI|JUNI|JULI|SEPTEMBER|OKTOBER|NOVEMBER|DESEMBER)\s+(\d{4})', line, re.IGNORECASE)
        if date_match:
            data['date_day'] = int(date_match.group(1))
            data['date_month_name'] = date_match.group(2).upper()
            data['date_year'] = int(date_match.group(3))
            continue

        if 'PEMASUKAN :' in line:
            current_section = 'pemasukan'
            continue
        elif 'NOTE BELUM BAYAR :' in line or 'SISA PEMBAYARAN :' in line:
            current_section = 'other'
            continue
        elif 'PENGELUARAN :' in line:
            current_section = 'pengeluaran'
            continue
        elif 'BELANJAAN KE LABURA:' in line:
            current_section = 'other'
            continue

        # Extract items
        if current_section == 'pemasukan':
            if 'TOTAL PEMASUKAN' in line or 'TOTAL TF' in line or 'TOTAL MANDIRI' in line:
                pass
            elif '=' in line and 'RB' in line:
                data['pemasukan_items'].append(line.replace('✅', '').strip())

        elif current_section == 'pengeluaran':
            if 'TOTAL PENGELUARAN' in line or 'TOTAL UANG' in line or 'SELISIH' in line:
                pass
            elif '=' in line and 'RB' in line:
                data['pengeluaran_items'].append(line.replace('✅', '').strip())

        # Extract Totals
        if 'TOTAL PEMASUKAN:' in line:
            m = re.search(r'([\d\.]+)\s*RB', line)
            if m:
                data['total_pemasukan'] = float(m.group(1).replace('.', ''))
        elif 'TOTAL TF BCA :' in line:
            m = re.search(r'([\d\.]+)\s*RB', line)
            if m:
                data['total_bca'] = float(m.group(1).replace('.', ''))
        elif 'TOTAL TF BNI :' in line:
            m = re.search(r'([\d\.]+)\s*RB', line)
            if m:
                data['total_bni'] = float(m.group(1).replace('.', ''))
        elif 'TOTAL TF BRI :' in line:
            m = re.search(r'([\d\.]+)\s*RB', line)
            if m:
                data['total_bri'] = float(m.group(1).replace('.', ''))
        elif 'TOTAL MANDIRI :' in line:
            m = re.search(r'([\d\.]+)\s*RB', line)
            if m:
                data['total_mandiri'] = float(m.group(1).replace('.', ''))
        elif 'TOTAL CASH :' in line:
            m = re.search(r'([\d\.]+)\s*RB', line)
            if m:
                data['total_cash'] = float(m.group(1).replace('.', ''))
        elif 'TOTAL PENGELUARAN:' in line:
            m = re.search(r'([\d\.]+)\s*RB', line)
            if m:
                data['total_pengeluaran'] = float(m.group(1).replace('.', ''))

    return data


def update_excel_native(excel_path, parsed_data):
    """Updates cell values in Excel file using openpyxl preserving 100% of formatting, colors, and layout."""
    wb = openpyxl.load_workbook(excel_path, keep_vba=True)

    sheet_name = parsed_data['date_month_name'] or 'AGUSTUS'
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet {sheet_name} not found in workbook!")

    sheet = wb[sheet_name]
    target_day = parsed_data['date_day'] or 10

    # Date header is at row 3 (1-indexed). Target column = target_day + 2
    # e.g., Day 1 = Col 3 (C), Day 10 = Col 12 (L)
    col_idx = target_day + 2
    col_letter = get_column_letter(col_idx)

    # 1. Pemasukan Summary Row = Row 4 (Leave Row 3 date intact!)
    sheet.cell(row=4, column=col_idx).value = parsed_data['total_pemasukan']

    # 2. Write Pemasukan items (Row 5 s/d Row 9)
    for i, item in enumerate(parsed_data['pemasukan_items']):
        row_idx = 5 + i
        if row_idx <= 13:
            sheet.cell(row=row_idx, column=col_idx).value = item

    # 3. Write Pemasukan Total (Row 14)
    sheet.cell(row=14, column=col_idx).value = f" Rp{parsed_data['total_pemasukan']:,.0f}.000 " if parsed_data['total_pemasukan'] else " Rp- "

    # 4. Write Bank Breakdowns (Rows 15-19)
    sheet.cell(row=16, column=col_idx).value = f" Rp{parsed_data['total_bni']:,.0f}.000 " if parsed_data['total_bni'] else None
    sheet.cell(row=17, column=col_idx).value = f" Rp{parsed_data['total_bca']:,.0f}.000 " if parsed_data['total_bca'] else None
    sheet.cell(row=18, column=col_idx).value = f" Rp{parsed_data['total_mandiri']:,.0f}.000 " if parsed_data['total_mandiri'] else None
    sheet.cell(row=19, column=col_idx).value = f" Rp{parsed_data['total_cash']:,.0f}.000 " if parsed_data['total_cash'] else None

    # 5. Write Total Pengeluaran (Row 22)
    sheet.cell(row=22, column=col_idx).value = f" Rp{parsed_data['total_pengeluaran']:,.0f}.000 " if parsed_data['total_pengeluaran'] else " Rp- "

    wb.save(excel_path)
    print(f"SUCCESS: openpyxl updated cell values in {sheet_name} (Col {col_letter}) preserving 100% visual formatting!")


if __name__ == '__main__':
    txt_file = 'e:/JS/laporan-test/REKAPAN TERBARU2.txt'
    excel_file = 'e:/JS/laporan-test/TABEL REKAPAN NEW2026-.xlsm'

    data = parse_txt_report(txt_file)
    print("Parsed Data from TXT:", data)
    update_excel_native(excel_file, data)
